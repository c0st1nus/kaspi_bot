from openpyxl import load_workbook, open, Workbook
from config import excel_old_path, excel_new_path, excel_price_path, x_tenge
from Bot import debug, send

def excel_actions(action, search_value=None, web_price=None, BotSend=None):

    if action == 'select':
        new_SKU = select_loop('new', 'A')
        new_price = select_loop('new', 'D')
        new_min_price = select_loop('new', 'K')

        old_id = select_loop('old', 'A')
        old_url = select_loop('old', 'G')

        return new_SKU, new_price, new_min_price, \
               old_id, old_url

    elif action == 'update':
        wb = load_workbook(filename=excel_new_path, data_only=True)
        sheet = wb.active

        # Цена конкурента - 1 тенге
        int_upd_price = web_price - x_tenge

        # int
        # target_value = "108602788_920758"

        number = 0
        # Перебираем ячейки в столбце A
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == search_value:
                    # Найдено совпадение, выводим номер строки
                    print(f'Номер строки с {search_value} в столбце A: {cell.row}')
                    number = cell.row
                    break


        # Цена (price) в new.xmlx
        new_price = sheet['D' + f'{number}']
        int_new_price = new_price.value

        # Минимальная цена (min price) в new.xmlx
        new_min_price = sheet['K'+f'{number}']

        int_new_min_price = int(new_min_price.value)

        # Цена конкурента
        int_website_price = int(web_price)


        if int_new_price != int_website_price > int_new_min_price < int_upd_price != int_new_price:

            # Изменяет в таблице
            # Новое значение po F + результат поиска (цифра)
            sheet['D' + f'{number}'] = int_upd_price
            wb.save(excel_new_path)
            debug(f'Kaspi-Bot обновил цену на {search_value}: \n Новая: {int_upd_price}, Старая: {int_new_price}, Min Price: {int_new_min_price}')
            print(f'Kaspi-Bot обновил цену: Новая: {int_upd_price}, Старая: {int_new_price}, Min Price: {int_new_min_price}')


        else:
            print('Kaspi-Bot: Обновление цены не требуется.')




        wb.close()


    elif action == 'cut':
        wb = load_workbook(filename=excel_new_path, data_only=True)
        sheet = wb.active

        sheet.delete_cols(11, 14)
        print('Kaspi-Bot: Таблица полностю обработана.')
        send()
        wb.save(excel_price_path)

# Достает столбцы
def select_loop(name, colum):
    if name == 'new':
        file_path = excel_new_path

    elif name == 'old':
        file_path = excel_old_path

    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb.active

    # Создаем пустой список для столбца
    column_values = []

    # Замените 'A' на букву столбца, который вы хотите получить
    column_letter = colum

    # Получаем все значения из столбца и добавляем их в список
    for cell in sheet[column_letter]:
        column_values.append(cell.value)

    return column_values


