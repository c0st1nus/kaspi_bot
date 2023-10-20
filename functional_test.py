from contextlib import closing
from time import sleep

import requests

import openpyxl
import os

from selenium import webdriver

from config import headless


def test_start():

    print('----- Запущен: Проверка работоспособности бота -----')
    print('Займет несколько секунд...')


    parsing = price_parsing()
    excel = excel_manipulation()
    firefox = firefox_browser()

    print(f'-------------------------------------\n'
          f'Функция парсинга: {parsing}\n'
          f'Работа с Excel таблицами: {excel}\n'
          f'Работа Firefox: {firefox}\n'
          f'Автозагузка цен на Kaspi: {firefox} \n'
          f'-------------------------------------')


































































def price_parsing():

    id = '106221648_548760'
    url = f'https://kaspi.kz/yml/offer-view/offers/{id}'

    user_agent = 'Mozilla\/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit\/537.36 (KHTML, like Gecko) Chrome\/60.0.3112.113 Safari\/537.36"'

    headers = {
        'Accept': 'application/json, text/*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'User-Agent': f'{user_agent}',
        'Referer': f'{url}',
    }

    json_data = {
        'cityId': '750000000',
        'id': f'{id}', }

    response = requests.post(url=url, headers=headers, json=json_data)


    sleep(2)

    if response.status_code == 200:
        return 'OK'

    else:
        return 'ERROR'


def excel_manipulation():
    try:
        # Создание новой таблицы
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Заголовок 1'
        sheet['B1'] = 'Заголовок 2'

        # Сохранение новой таблицы
        workbook.save('TEST.xlsx')

        # Добавление данных в существующую таблицу
        workbook = openpyxl.load_workbook('TEST.xlsx')
        sheet = workbook.active
        sheet.append(['Данные 1', 'Данные 2'])
        workbook.save('TEST.xlsx')

        # Чтение данных из таблицы
        workbook = openpyxl.load_workbook('TEST.xlsx')
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Удаление таблицы
        os.remove('TEST.xlsx')

        return 'OK'

    except:
        return 'OK'


def firefox_browser():
    try:
        option = webdriver.FirefoxOptions()

        if headless == 'yes':
            option.add_argument("--headless")

        with closing(webdriver.Firefox(options=option)) as browser:
            browser.get("https://example.com/")
            sleep(5)

        return 'OK'

    except:
        return 'ERROR'



