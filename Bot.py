import json
import random
import telebot
from telebot import types
from telebot.types import InlineKeyboardMarkup
from pathlib import Path
import openpyxl
from UsersData.Processing import loop
import UsersData.browser_bot
import threading
from time import sleep
import os

token = '6932808440:AAHZF-iXcPhY58CFZ0J3d3pUdgt7je7n7S4'

bot = telebot.TeleBot(token)
admin = 1239398217
def my_background_function():
    while True:
        loop()


background_thread = threading.Thread(target=my_background_function)
background_thread.daemon = True
background_thread.start()


@bot.message_handler(commands=['start'])
def start(message):
    with open('UsersData/Users.json', 'r') as file:
        try:
            data = json.load(file)
        except:
            data = None
    if message.from_user.id != admin:
        if message.from_user.id in data['loginedUserID']:  # проверка на то, что пользователь ранее уже входил в аккаунт
            path = str(data["usernames"][data['loginedUserID'].index(message.from_user.id)])
            if not Path(f'UsersData/{path}/config.json').is_file():
                with open(f'UsersData/{path}/config.json', 'w+') as file:
                    data2 = {
                        'Login': None,
                        'Pass': None,
                        'excel_new_path': f'UsersData/{path}/new.xlsx',
                        'excel_old_path': f'UsersData/{path}/old.xlsx',
                        'excel_price_path': f'UsersData/{path}/price_new.xlsx'
                    }
                    json.dump(data2, file)
                if not Path(f'UsersData/{path}/new.xlsx').is_file(): # проверка наличие таблиц
                    workbook = openpyxl.Workbook()
                    workbook.save(f'UsersData/{path}/new.xlsx')
                if not Path(f'UsersData/{path}/old.xlsx').is_file():
                    workbook = openpyxl.Workbook()
                    workbook.save(f'UsersData/{path}/old.xlsx')
                if not Path(f'UsersData/{path}/price_new.xlsx').is_file():
                    workbook = openpyxl.Workbook()
                    workbook.save(f'UsersData/{path}/price_new.xlsx')       # создаем config.json и 3 эксель файла если их нет #создаем config.json и 3 эксель файла если их нет   #создаем config.json и 3 эксель файла если их нет
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True) # вывод стартогого меню
            keyboard.add(types.KeyboardButton('Выйти'), types.KeyboardButton('Добавить данные'),
                         types.KeyboardButton('Мои товары'), types.KeyboardButton('Как пользоваться ботом?'), types.KeyboardButton('Запустить/остановить бот'))
            bot.send_message(message.chat.id, 'Вы зашли в главное меню', reply_markup=keyboard)
            index = data['loginedUserID'].index(message.from_user.id)
            bot.register_next_step_handler(message, menu, index)
        else:
            keyboard: InlineKeyboardMarkup = types.InlineKeyboardMarkup() # регистрации/авторизация
            keyboard.add(types.InlineKeyboardButton('Авторизироваться', callback_data='Log'))
            bot.send_message(message.chat.id, "Здравствуйте! У вас есть аккаунт или вы здесь впервые?",
                                 reply_markup=keyboard)
    else:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.row(types.KeyboardButton('Добавить пользователя'), types.KeyboardButton('Удалить пользователя'), types.KeyboardButton('Остановить/Запустить бот для пользователя'))
        keyboard.row(types.KeyboardButton('Список пользователей'), types.KeyboardButton('Добавить данные'), types.KeyboardButton('Мои товары'))
        if not Path('UsersData/admin/config.json').is_file():
            with open('UsersData/admin/config.json', 'w+') as file:
                data2 = {
                    'Login': None,
                    'Pass': None,
                    'excel_new_path': 'UsersData/admin/new.xlsx',
                    'excel_old_path': 'UsersData/admin/old.xlsx',
                    'excel_price_path': 'UsersData/admin/price_new.xlsx'
                }
                json.dump(data2, file)
            if not Path(f'UsersData/admin/new.xlsx').is_file():  # проверка наличие таблиц
                workbook = openpyxl.Workbook()
                workbook.save(f'UsersData/admin/new.xlsx')
            if not Path(f'UsersData/admin/old.xlsx').is_file():
                workbook = openpyxl.Workbook()
                workbook.save(f'UsersData/admin/old.xlsx')
            if not Path(f'UsersData/admin/price_new.xlsx').is_file():
                workbook = openpyxl.Workbook()
                workbook.save(f'UsersData/admin/price_new.xlsx')
        bot.send_message(message.chat.id, 'Вы зашли в главное меню', reply_markup=keyboard)
        bot.register_next_step_handler(message, menu, 0)


def productDelete(message, index): # удаления данных продукта из таблицы new.xlsx
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    with open(f'UsersData/{data["usernames"][index]}/config.json', 'r') as file:
        userdata = json.load(file)
    delete_row_by_value(userdata['excel_new_path'], message.text)
    bot.send_message(message.chat.id, 'Товар успешно удален')
    sleep(1)
    start(message)

@bot.message_handler(content_types=['document'])
def productAdd(message, index): #добавление данных товара в таблицу new.xlsx
    try:
        with open('UsersData/Users.json', 'r') as file:
            data = json.load(file)
        with open(f'UsersData/{data["usernames"][index]}/config.json', 'r') as file:
            userdata = json.load(file)
        old_path = userdata["excel_old_path"]
        new_path = userdata["excel_new_path"]
        file_info = bot.get_file(message.document.file_id)
        file = bot.download_file(file_info.file_path)
        with open(message.document.file_name, 'wb') as new_file:
            new_file.write(file)
        if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            bot.send_message(message.chat.id, 'Бот добавляет ваши товары...')
            source_wb = openpyxl.load_workbook(message.document.file_name)
            source_sheet = source_wb.active
            new = openpyxl.load_workbook(new_path)
            new_sheet = new.active
            old = openpyxl.load_workbook(old_path)
            old_sheet = old.active
            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)
                old_sheet.append(row)
                new.save(f'UsersData/{data["usernames"][index]}/new.xlsx')
                old.save(f'UsersData/{data["usernames"][index]}/old.xlsx')
            source_wb.close()
            os.remove(message.document.file_name)
            bot.send_message(message.chat.id, 'Данные успешно сохранены')
            if userdata['Login']:
                with open('UsersData/ReadyUsers.json', 'r') as file:
                    users = json.load(file)
                users['usernames'].append(data["usernames"][index])
                with open('UsersData/ReadyUsers.json', 'w') as file:
                    json.dump(users, file)
                bot.send_message(message.chat.id, 'Загрузка данных окончена, бот начал работу')
        else:
            bot.send_message(message.chat.id, 'Вы отправили файл некоректного типа')
    except:
        bot.send_message(message.chat.id, 'Вы совершили ошибку при загрузке товара')
    sleep(1)
    start(message)
    

def product(message, index): # меню "мои товары"
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    with open(f'UsersData/{data["usernames"][index]}/config.json', 'r') as file:
        userdata = json.load(file)
    if str(message.text) == 'Список товаров':
        if count_filled_cells(userdata['excel_new_path']) == []:
            bot.send_message(message.chat.id, 'Вы не еще заполнили не одни товар')
        else:
            answer = ''
            result = count_filled_cells(userdata['excel_new_path'])
            for i in range(len(result)):
                if i != 0:
                    answer += str(result[i]) + '\n'
            bot.send_message(message.chat.id, text=answer)
    elif message.text == 'Удалить товар':
        if count_filled_cells(userdata['excel_new_path']) == []:
            bot.send_message(message.chat.id, 'Вы не еще заполнили не одни товар')
            sleep(1)
            start(message)
        else:
            bot.send_message(message.chat.id, 'Напшите артикул товара, который хотите удалить')
            bot.register_next_step_handler(message, productDelete, index)
    elif message.text == 'Добавить товар':
        bot.send_message(message.chat.id, 'Отправьте exel таблицу с вашими данными')
        bot.register_next_step_handler(message, productAdd, index)


def dataAdd(message, index): # добавление логина и пароля от Каспия
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    with open(f'UsersData/{data["usernames"][index]}/config.json', 'r') as file:
        userdata = json.load(file)
    data2 = message.text.split(', ')
    if len(data2) != 2:
        bot.send_message(message.chat.id,
                         'Вы допустили ошибку при вводе данных, логин и пароль должны быть в одном сообщении через запятную(логин, пароль)')
        sleep(1)
        start(message)
    else:
        bot.send_message(message.chat.id, 'Бот пытается войти в ваш акаунт')
        if UsersData.browser_bot.try_to_sign_in_upload_kaspi(data2[0], data2[1]):
            bot.send_message(message.chat.id, 'Успешно, данные сохранены')
            userdata['Login'] = data2[0]
            userdata['Pass'] = data2[1]
            with open(f'UsersData/{data["usernames"][index]}/config.json', 'w') as file:
                json.dump(userdata, file)
            if count_filled_cells(userdata['excel_new_path']):
                with open('UsersData/ReadyUsers.json', 'r') as file:
                    users = json.load(file)
                users['usernames'].append(data["usernames"][index])
                with open('UsersData/ReadyUsers.json', 'w') as file:
                    json.dump(users, file)
                bot.send_message(message, 'Загрузка данных окончена, бот начал работу')
            start(message)
        else:
            bot.send_message(message.chat.id, 'Произошла ошибка, попробуйте еще раз')
            bot.send_message(message.chat.id,
                             'Введите логин и пароль от https://kaspi.kz/mc в ОДНОМ сообщении через запятую(логин, пароль)')
            bot.register_next_step_handler(message, dataAdd, index)
def stop_start(message):
    if message.from_user.id == admin:
        with open('UsersData/Users.json', 'r') as file:
            Fulldata = json.load(file)
        if message.text in Fulldata['usernames']:
            with open('UsersData/ReadyUsers.json', 'r') as file:
                data = json.load(file)
            if message.text in data['usernames']:
                index = data['usernames'].index(message.text)
                data['usernames'].pop(index)
                bot.send_message(message.chat.id, f'Бот для остановлен для пользователя {message.text}')
            else:
                with open(f'UsersData/{message.text}/config.json', 'r') as file:
                    userData = json.load(file)
                if count_filled_cells(userData['excel_new_path']) != [] and userData['Login'] != None:
                    data['usernames'].append(message.text)
                    bot.send_message(message.chat.id, f'Бот для запущен для пользователя {message.text}')
            with open('UsersData/ReadyUsers.json', 'w') as file:
                json.dump(data, file)
            sleep(1)
            start(message)
        else:
            bot.send_message(message.chat.id, 'Данный пользователь не существует')
            sleep(1)
            start(message)
    else:
        with open('UsersData/Users.json', 'r') as file:
            data = json.load(file)
        index = data['loginedUserID'].index(message.from_user.id)
        with open(f'UsersData/{data["usernames"][index]}/config.json') as file:
            userData = json.load(file)
        with open('UsersData/ReadyUsers.json', 'r') as file:
            ready = json.load(file)
        if data["usernames"][index] in ready:
            ready['usernames'].pop(index)
            bot.send_message(message.chat.id, f'Бот остановлен')
        else:
            if count_filled_cells(userData['excel_new_path']) and userData['Login'] != None:
                ready["usernames"].append(data["usernames"][index])
                bot.send_message(message.chat.id, f'Бот запущен')
            else:
                bot.send_message(message.chat.id, f'У вас не до конца заполнены данные')
        with open('UsersData/ReadyUsers.json', 'w') as file:
            json.dump(ready, file)
        sleep(1)
        start(message)

def menu(message, index=None):  # основное меню бота
    try:
        with open('UsersData/Users.json', 'r') as file:
            try:
                data = json.load(file)
            except:
                data = None
        if message.text == '/start':
            bot.register_next_step_handler(message, start)
        elif message.text == 'Запустить/остановить бот':
            stop_start(message)
        elif message.text == 'Добавить данные':
            with open(f'UsersData/{data["usernames"][index]}/config.json', 'r') as file:
                userdata = json.load(file)
            if userdata['Login'] == None:
                bot.send_message(message.chat.id,
                                 'Введите логин и пароль от https://kaspi.kz/mc в ОДНОМ сообщении через запятую(логин, пароль)')
                bot.register_next_step_handler(message, dataAdd, index)
            else:
                global s
                s = index
                keyboard = types.InlineKeyboardMarkup()
                keyboard.add(types.InlineKeyboardButton('Да', callback_data='changeKaspiData'),
                             types.InlineKeyboardButton('Нет, вернуться в меню', callback_data='ReturnToMainMenu'))
                bot.send_message(message.chat.id, 'У вас уже введены данные от Kaspi, вы хотите их изменить?1',
                                 reply_markup=keyboard)
        elif message.text == 'Мои товары':
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.row(types.KeyboardButton('Список товаров'))
            keyboard.row(types.KeyboardButton('Добавить товар'), types.KeyboardButton('Удалить товар'))
            bot.send_message(message.chat.id, 'Меню "Мои товары"', reply_markup=keyboard)
            bot.register_next_step_handler(message, product, index)
        elif message.text == 'Остановить/Запустить бот для пользователя':
            bot.send_message(message.chat.id, 'Напишите логин пользователя')
            bot.register_next_step_handler(message, stop_start)
        elif message.text == 'Как пользоваться ботом?':
            with open('Files/instruction.txt', 'rb') as file:
                text = file.read()
            photo = open('Files/picture.jpg', 'rb')
            bot.send_photo(message.chat.id, photo, text)
            sleep(5)
            start(message)
        elif message.text == 'Список пользователей':
            if data:
                text = ''
                for i in range(len(data['usernames'])):
                    text += data['usernames'][i] + ':' + str(data['passwords'][i]) + '\n'
                bot.send_message(message.chat.id, text)
            else:
                bot.send_message(message.chat.id, 'Еще не зарегистрированно ни одного пользователя')
            sleep(3)
            start(message)
        elif message.text == 'Выйти':
            data['loginedUserID'][data['loginedUserID'].index(f'{message.from_user.id}')] = None
            bot.send_message(message.chat.id, 'Выход выполнен')
            with open('UsersData/Users.json', 'w') as file:
                json.dump(data, file)
        elif message.text == 'Добавить пользователя':
            bot.send_message(message.chat.id, 'Введите логин пользователя')
            bot.register_next_step_handler(message, reg1)
        elif message.text == 'Удалить пользователя':
            bot.send_message(message.chat.id, 'Введите логин пользователя')
            bot.register_next_step_handler(message, deleteUser)
    except:
        bot.send_message(message.chat.id, 'Возникла ошибка, обратитесь к администратору')
def deleteUser(message):
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    index = data['usernames'].index(message.text)
    data['usernames'].pop(index)
    data['passwords'].pop(index)
    try:
        data['loginedUserID'].pop(index)
    except:
        pass
    with open('UsersData/Users.json', 'w') as file:
        json.dump(data, file)
    bot.send_message(message.chat.id, f'Пользователь {message.text} успешно удален')
    sleep(1)
    start(message)
def reg1(message):
    login = message.text
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(types.KeyboardButton('Сгенерируй пароль'))
    bot.send_message(message.chat.id, 'Отлично, теперь введите пароль пользователя', reply_markup=keyboard)
    bot.register_next_step_handler(message, reg2, login)

def reg2(message, login):
    password = None
    if message.text == 'Сгенерируй пароль':
        password = randompassword(8)
    else:
        password = message.text
    try:
        with open('UsersData/Users.json', 'r') as file:
            data = json.load(file)
        data['usernames'].append(login)
        data['passwords'].append(password)
        with open('UsersData/Users.json', 'w') as file:
            json.dump(data, file)
    except:
        data = {"usernames": [], "passwords": [], "loginedUserID": []}
        data['usernames'].insert(0, login)
        data['passwords'].insert(0, password)
        with open('UsersData/Users.json', 'w') as file:
            json.dump(data, file)
    os.mkdir(f"UsersData/{login}")
    bot.send_message(message.chat.id, f'Данные успешно сохранены, пароль пользователя {login} - {password}')
    sleep(1)
    start(message)

def login1(message):  # поиск логина в Users.json
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    if message.text == '/start':
        bot.register_next_step_handler(message, start)
    elif message.text.lower() in data['usernames']:
        bot.send_message(message.chat.id, 'Введите ваш пароль')
        index = data['usernames'].index(message.text.lower())
        bot.register_next_step_handler(message, login2, index)
    else:
        bot.send_message(message.chat.id, 'Вы ввели не верный логин')
        bot.send_message(message.chat.id, 'Введите ваш логин')
        bot.register_next_step_handler(message, login1)


def login2(message, index):  # проверка пароля и сохранение айди пользователя в Users.json при положительном варианте
    with open('UsersData/Users.json', 'r') as file:
        data = json.load(file)
    if message.text == '/start':
        bot.register_next_step_handler(message, start)
    elif message.text == data['passwords'][index]:
        try:
            data['loginedUserID'][index] = message.from_user.id
            with open('UsersData/Users.json', 'w') as file:
                json.dump(data, file)
        except:
            data['loginedUserID'].append(message.from_user.id)
            with open('UsersData/Users.json', 'w') as file:
                json.dump(data, file)
        bot.send_message(message.chat.id, 'Вход выполнен')
        sleep(1)
        start(message)
        

@bot.callback_query_handler(func=lambda callback: True)
def callback_message(callback): # обработка нажатия Inline кнопок
    if callback.data == 'Log':
        bot.send_message(callback.message.chat.id, 'Введите ваш логин')
        bot.register_next_step_handler(callback.message, login1)
    elif callback.data == 'changeKaspiData':
        global s
        with open('UsersData/Users.json', 'r') as file:
            data = json.load(file)
        with open(f'UsersData/{data["usernames"][s]}/config.json', 'r') as file:
            userdata = json.load(file)
        userdata['Login'] = None
        userdata['Pass'] = None
        with open(f'UsersData/{data["usernames"][s]}/config.json', 'w') as file:
            json.dump(userdata, file)
        print(s)
        bot.send_message(callback.message.chat.id,
                         'Введите логин и пароль от https://kaspi.kz/mc в ОДНОМ сообщении через запятую(логин, пароль)')
        bot.register_next_step_handler(callback.message, dataAdd, s)
    elif callback.data == 'ReturnToMainMenu':
        sleep(1)
        start(callback.message)


def randompassword(count):  # генерация пароля
    randm = []
    for i in range(count):
        randm.append(str(random.randint(0, 10)))
    sep = ''
    return sep.join(randm)


def articles(excel_path, sheetToAnalyze): # функция выводящая список данных из задоного столбца
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    column_data_a = sheet[sheetToAnalyze]
    filled_cells = []
    for cell_a in column_data_a:
        if cell_a.value is not None:
            filled_cells.append(str(cell_a.value))
    if len(filled_cells) == 1:
        return []
    else:
        return (filled_cells)


def count_filled_cells(excel_path): # вывод тайтла т цены товара
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    column_data_a = sheet['A']
    column_data_b = sheet['B']
    filled_cells = []

    for cell_a, cell_b in zip(column_data_a, column_data_b):
        if cell_a.value is not None and cell_b.value is not None:
            filled_cells.append(f'{cell_a.value}: {cell_b.value}')
    if len(filled_cells) == 1:
        return []
    else:
        return (filled_cells)

def delete_row_by_value(excel_path, value_to_delete): # удаление данных из таблицы
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == value_to_delete:
                sheet.delete_rows(cell.row)
                break
    workbook.save(excel_path)
    workbook.close()


if __name__ == "__main__":
    bot.polling(none_stop=True)
