import time
import telebot
from telebot import types
from pathlib import Path
from openpyxl import load_workbook, Workbook
import os
from functools import wraps
from UsersData.handler import DatabaseHandler, db_connection
from UsersData.browser_bot import try_to_sign_in_upload_kaspi
from UsersData.Processing import loop
import threading
from time import sleep

token = "TOKEN"
bot = telebot.TeleBot(token)

animate = True
wait_smile = ['🕐','🕑','🕒','🕓','🕔','🕕','🕖','🕗','🕘','🕙','🕚','🕛']

def my_background_function():
    while True:
        loop()


background_thread = threading.Thread(target=my_background_function)
background_thread.daemon = True
background_thread.start()

def handle_errors(func):
    @wraps(func)
    def wrapper(message, *args, **kwargs):
        try:
            return func(message, *args, **kwargs)
        except Exception as e:
            bot.send_message(message.chat.id, '⚠️ Произошла ошибка. Пожалуйста, попробуйте снова или обратитесь к администратору.')
            print(f"Error in {func.__name__}: {str(e)}")
    return wrapper

def create_main_keyboard():
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.KeyboardButton(btn) for btn in [
        'Добавить данные', 'Мои товары', 
        'Как пользоваться ботом?', 'Запустить/остановить бот'
    ]])
    return keyboard

@bot.message_handler(commands=['start'])
@handle_errors
@db_connection
def start(message, conn):
    if not conn.query_data("SELECT * FROM users WHERE telegram_id = ?", (message.from_user.id,)):
        conn.insert_data('users', {'telegram_id': message.from_user.id})

    # Create user directory structure
    user_dir = Path(f'UsersData/{message.from_user.id}')
    user_dir.mkdir(exist_ok=True)
    
    # Create default Excel file if not exists
    excel_file = user_dir / 'pricelist.xlsx'

    if not excel_file.exists():
        workbook = Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1, value="article")
        sheet.cell(row=1, column=2, value="description")
        sheet.cell(row=1, column=4, value="price")
        sheet.cell(row=1, column=5, value="PP1")
        sheet.cell(row=1, column=11, value="min_price")

        workbook.save(excel_file)
        workbook.close()
        

    # Create main keyboard
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.KeyboardButton(btn) for btn in [
        'Добавить данные', 'Мои товары', 
        'Как пользоваться ботом?', 'Запустить/остановить бот'
    ]])
    
    bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                    reply_markup=create_main_keyboard())
    bot.register_next_step_handler(message, menu)

@handle_errors
@db_connection
def menu(message, conn):
    if message.text == 'Запустить/остановить бот':
        handle_bot_status(message)
    elif message.text == 'Добавить данные':
        handle_add_data(message)
    elif message.text == 'Мои товары':
        show_products_menu(message)
    elif message.text == 'Как пользоваться ботом?':
        send_instructions(message)

@handle_errors
@db_connection
def handle_bot_status(message, conn):
    user_id = message.from_user.id
    login = conn.query_data("SELECT login FROM users WHERE telegram_id = ?", (user_id,))
    products = conn.query_data("SELECT * FROM products WHERE user_id = ?", (user_id,))
    
    if not login or not products:
        bot.send_message(message.chat.id, '❌ Для запуска бота необходимо добавить данные Каспи и товары')
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)
    
    status = conn.query_data("SELECT bot_status FROM users WHERE telegram_id = ?", (user_id,))[0]['bot_status']
    new_status = not status
    
    conn.update_data('users', {'bot_status': new_status}, "telegram_id = ?", (user_id,))
    bot.send_message(message.chat.id, f'🤖 Бот {"запущен" if new_status else "остановлен"}')
    bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
    bot.register_next_step_handler(message, menu)

@handle_errors
@db_connection
def handle_add_data(message, conn):
    user_data = conn.query_data("SELECT login, password FROM users WHERE telegram_id = ?", (message.from_user.id,))
    if user_data[0]["login"] and user_data[0]["password"]:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton('Да', callback_data='changeKaspiData'),
            types.InlineKeyboardButton('Нет', callback_data='ReturnToMainMenu')
        )
        bot.send_message(message.chat.id, '⚙️ У вас уже есть сохраненные данные. Хотите изменить их?', reply_markup=keyboard)
    else:
        bot.send_message(message.chat.id, '📝 Введите логин и пароль от Kaspi через запятую (формат: логин, пароль)')
        bot.register_next_step_handler(message, process_credentials)

def animate_smiles(message):
    i = 0
    while animate:
        bot.edit_message_text(
            f'{wait_smile[i]} Бот пытается войти в ваш акаунт', 
            message.chat.id, 
            message.message_id
        )
        i = (i + 1) % 12
        time.sleep(0.25)

@db_connection
@handle_errors
def process_credentials(message, conn: DatabaseHandler):
    global animate
    if len(message.text.split(', ')) != 2:
        bot.send_message(message.chat.id, '❌ Неверный формат данных. Пожалуйста, используйте формат: логин, пароль')
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)
    else:
        login, password = message.text.split(', ')
        s_message = bot.send_message(message.chat.id, 'Бот пытается войти в ваш акаунт')
        animation_thread = threading.Thread(target=animate_smiles, args=(s_message,))
        animation_thread.start()

        login_result = try_to_sign_in_upload_kaspi(login, password)

        animate = False
        animation_thread.join()

        if login_result:
            conn.update_data('users', {'login': login, 'password': password}, "telegram_id = ?", (message.from_user.id,))
            bot.send_message(message.chat.id, '🔐 Данные успешно сохранены')
        else:
            bot.send_message(message.chat.id, '❌ Неверный логин или пароль')
        time.sleep(0.5)
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)

@handle_errors
def show_products_menu(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row('Список товаров')
    keyboard.row('Добавить товар', 'Удалить товар')
    bot.send_message(message.chat.id, '📦 Меню товаров:', reply_markup=keyboard)
    bot.register_next_step_handler(message, handle_products)

@handle_errors
@db_connection
def handle_products(message, conn):
    user_id = message.from_user.id
    if message.text == 'Список товаров':
        products = conn.query_data("SELECT * FROM products WHERE user_id = ?", (user_id,))
        if not products:
            bot.send_message(message.chat.id, '📭 У вас пока нет товаров')
        else:
            response = '\n'.join([f"{idx+1}. {p['description']} - {p["price"]} - (арт. {p['article']})" 
                                for idx, p in enumerate(products)])
            bot.send_message(message.chat.id, f'📦 Ваши товары:\n{response}')
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)
    elif message.text == 'Добавить товар':
        bot.send_message(message.chat.id, '📎 Отправьте Excel-файл с данными товаров')
        bot.register_next_step_handler(message, process_excel_file)
    elif message.text == 'Удалить товар':
        bot.send_message(message.chat.id, '✏️ Введите артикул товара для удаления')
        bot.register_next_step_handler(message, process_product_deletion)
    else:
        bot.send_message(message.chat.id, "⚠️ Неизвестная команда в меню товаров")
        show_products_menu(message)

@handle_errors
@db_connection
def process_excel_file(message, conn: DatabaseHandler):
    try:
        file_info = bot.get_file(message.document.file_id)
        file = bot.download_file(file_info.file_path)
        
        # Save with unique filename
        filename = f"temp_{message.from_user.id}.xlsx"
        with open(filename, 'wb') as f:
            f.write(file)
            
        wb = load_workbook(filename)
        sheet = wb.active
        
        # Process rows (example implementation)
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            existing_product = conn.query_data("SELECT * FROM products WHERE article = ?", (row[0],))
            price = row[3]
            min_price = row[10] if row[10] else 0
    
            if existing_product:
                conn.update_data('products', {
                    'user_id': message.from_user.id,
                    'article': row[0],    # A
                    'description': row[1],# B
                    'shop_article': row[2], # C
                    'price': price, # D
                    'minimal_price': min_price # K
                }, "article = ?", (row[0],))
            else:
                conn.insert_data('products', {
                    'user_id': message.from_user.id,
                    'article': row[0],    # A
                    'description': row[1],# B
                    'shop_article': row[2], # C
                    'price': price, # D
                    'minimal_price': min_price # K
                })
        wb.save(filename)
        wb.close()
        os.remove(filename)
        bot.send_message(message.chat.id, '✅ Товары успешно добавлены')
    except Exception as e:
        bot.send_message(message.chat.id, '❌ Ошибка обработки файла')
        print(f"Excel processing error: {str(e)}")
    finally:
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)


@handle_errors
def send_instructions(message):
    try:
        instruction_path = Path('files/instruction.txt')
        picture_path = Path('files/picture.jpg')
        
        with open(instruction_path, 'r', encoding='utf-8') as file:
            text = file.read()
        
        with open(picture_path, 'rb') as photo:
            bot.send_photo(message.chat.id, photo, caption=text[:1024])  # Truncate to 1024 chars for caption
        
    except Exception as e:
        print(f"Error sending instructions: {str(e)}")
        bot.send_message(message.chat.id, '❌ Не удалось загрузить инструкцию')
    finally:
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)

@handle_errors
@db_connection
def process_product_deletion(message, conn):
    user_id = message.from_user.id
    article = message.text.strip()
    
    try:
        deleted = conn.delete_data(
            'products', 
            "user_id = ? AND article = ?", 
            (user_id, article)
        )
        
        if deleted:
            bot.send_message(message.chat.id, f'✅ Товар с артикулом {article} успешно удален')
        else:
            bot.send_message(message.chat.id, f'❌ Товар с артикулом {article} не найден')
            
    except Exception as e:
        print(f"Delete error: {str(e)}")
        bot.send_message(message.chat.id, '❌ Ошибка при удалении товара')
    finally:
        bot.send_message(message.chat.id, '🏠 Вы в главном меню', 
                        reply_markup=create_main_keyboard())
        bot.register_next_step_handler(message, menu)   

@bot.callback_query_handler(func=lambda callback: True)
def handle_callbacks(callback):
    if callback.data == 'changeKaspiData':
        bot.send_message(callback.message.chat.id, '📝 Введите новые данные через запятую (логин, пароль)')
        bot.register_next_step_handler(callback.message, process_credentials)
    elif callback.data == 'ReturnToMainMenu':
        start(callback.message)



if __name__ == "__main__":
    bot.polling(none_stop=True)