from openpyxl import load_workbook, Workbook
excel_new_path = None
excel_old_path = None
excel_price_path = None
def excel_actions(action, search_value=None, web_price=None, username=None):
    global excel_new_path, excel_old_path, excel_price_path
    excel_new_path = f'UsersData/{username}/new.xlsx'
    excel_old_path = f'UsersData/{username}/old.xlsx'
    excel_price_path = f'UsersData/{username}/price_new.xlsx'
    if action == 'select':
        new_SKU = select_loop('new', 'A')
        new_price = select_loop('new', 'D')
        new_min_price = select_loop('new', 'K')

        old_id = select_loop('old', 'A')
        old_url = select_loop('old', 'G')

        return new_SKU, new_price, new_min_price, \
               old_id, old_url

    elif action == 'update':
        wb = load_workbook(filename=excel_new_path, data_only=True)
        sheet = wb.active

        # Цена конкурента - 1 тенге
        int_upd_price = int(web_price) - 1

        # int
        # target_value = "108602788_920758"

        number = 0
        # Перебираем ячейки в столбце A
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == search_value:
                    # Найдено совпадение, выводим номер строки
                    print(f'Номер строки с {search_value} в столбце A: {cell.row}')
                    number = cell.row
                    break


        # Цена (price) в new.xmlx
        new_price = sheet['D' + f'{number}']
        int_new_price = new_price.value

        # Минимальная цена (min price) в new.xmlx
        new_min_price = sheet['K'+f'{number}']

        int_new_min_price = int(new_min_price.value)

        # Цена конкурента
        int_website_price = int(web_price)


        if int_new_price != int_website_price > int_new_min_price < int_upd_price != int_new_price:

            # Изменяет в таблице
            # Новое значение po F + результат поиска (цифра)
            sheet['D' + f'{number}'] = int_upd_price
            print(f'Kaspi-Bot обновил цену: Новая: {int_upd_price}, Старая: {int_new_price}, Min Price: {int_new_min_price}')


        else:
            print('Kaspi-Bot: Обновление цены не требуется.')
        wb.save(excel_new_path)



    elif action == 'cut':
        wb = load_workbook(filename=excel_new_path, data_only=True)
        sheet = wb.active
        sheet.delete_cols(11, 14)
        updated_values = []
        i = 2
        while i <= sheet.max_row:
            cell = sheet.cell(row=i, column=4)
            try:
                if sheet.cell(row=i, column=1).value not in updated_values:
                    cell.value = int(cell.value)
                    updated_values.append(sheet.cell(row=i, column=1).value)
                    i += 1
                else:
                    sheet.delete_rows(i, amount=1)
            except:
                sheet.delete_rows(i, amount=1)
        wb.save(excel_price_path)
        print('Kaspi-Bot: Таблица полностью обработана.')

# Достает столбцы
def select_loop(name, colum):
    global excel_new_path, excel_old_path, excel_price_path
    if name == 'new':
        file_path = excel_new_path

    elif name == 'old':
        file_path = excel_old_path

    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb.active

    # Создаем пустой список для столбца
    column_values = []

    # Замените 'A' на букву столбца, который вы хотите получить
    column_letter = colum

    # Получаем все значения из столбца и добавляем их в список
    for cell in sheet[column_letter]:
        column_values.append(cell.value)

    return column_values

