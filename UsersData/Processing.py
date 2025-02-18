import random
from time import sleep
from user_agents.random_agnet import get_agent
from UsersData.browser_bot import sign_in_upload_kaspi, get_previous_upload_date
from UsersData.handler import DatabaseHandler, db_connection
from kaspi.get import request
from openpyxl import load_workbook, Workbook

@db_connection
def loop(conn: DatabaseHandler):
    conn.create_table("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id VARCHAR(20) UNIQUE NOT NULL,
            login VARCHAR(50),
            password VARCHAR(50),
            bot_status BOOLEAN DEFAULT 0
        )""")
    
    conn.create_table("""
        CREATE TABLE IF NOT EXISTS products (
            article VARCHAR(50) NOT NULL PRIMARY KEY,
            user_id VARCHAR(20) NOT NULL,
            model VARCHAR(50) NOT NULL,
            brand VARCHAR(50) NOT NULL,
            shop_article VARCHAR(50) NOT NULL,
            price INTEGER NOT NULL,
            minimal_price INTEGER NOT NULL,
            PP1 BOOLEAN DEFAULT 0,
            PP2 BOOLEAN DEFAULT 0,
            PP3 BOOLEAN DEFAULT 0,
            PP4 BOOLEAN DEFAULT 0,
            PP5 BOOLEAN DEFAULT 0,
            preorders  DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users (telegram_id)
        )""")
    users = conn.query_data("SELECT telegram_id FROM users WHERE bot_status = 1")
    if not users:
        print('Нет активных пользователей')
        print('===============================================')
        randm = random.randint(30, 45)
        print(f'Kaspi-Bot: Следующий старт через: {randm} сек.')
        sleep(randm)
    else:
        for user in users:
            workbook = load_workbook(f"UsersData/{user['telegram_id']}/pricelist.xlsx")
            sheet = workbook.active
            sheet.delete_rows(2, sheet.max_row)
            workbook.save(f"UsersData/{user['telegram_id']}/pricelist.xlsx")
            workbook.close()
            print(f'Пользователь: {user["telegram_id"]}')
            print('===============================================')
            
            basic_bot_operation(user['telegram_id'])
            sign_in_upload_kaspi(user['telegram_id'])
            sleep(1)
        print(f'Kaspi-Bot: Следующий старт через: 30 мин.')
        sleep(1800)

@db_connection
def basic_bot_operation(telegram_id, conn: DatabaseHandler):
    products = conn.query_data("SELECT * FROM products WHERE user_id = ?", (telegram_id,))
    
    if not products:
        print(f'Kaspi-Bot: Нет товаров для пользователя {telegram_id}')
        return
    
    print(f'Kaspi-Bot: Будет обработано {len(products)} товаров.')
    total_products = len(products)
    
    for index, product in enumerate(products, 1):
        print('===============================================')
        print(f'Цикл: {index} / {total_products}')
        print(product)
        handler(product)
        randm = random.randint(1, 3)
        print(f'Kaspi-Bot: Ожидаем: {randm} сек.')
        sleep(randm)

@db_connection
def handler(product, conn: DatabaseHandler):
    global_user_agent = get_agent()
    result = request(id_id=product["shop_article"], user_agent=global_user_agent)

    if result == 'error':
        print('Kaspi-Bot: Не удалось получить цену конкурентов.')
        sleep(1)
        return

    try:
        current_price = int(result["price"])
    except ValueError:
        print('Kaspi-Bot: Некорректный формат цены')
        return

    web_price = current_price - 1
    stored_price = int(product["price"])
    min_price = int(product["minimal_price"])

    file_path = f"UsersData/{product['user_id']}/pricelist.xlsx"
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
    if web_price < min_price:
        web_price = min_price
    if web_price == stored_price:
        print(f'Kaspi-Bot: Цена не изменилась: {web_price}')
        return
    else:
        conn.update_data("products", {"price": web_price}, "article = ?", (product["article"],))
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=product["article"])
        sheet.cell(row=next_row, column=2, value=product["model"])
        sheet.cell(row=next_row, column=3, value=product["brand"])
        sheet.cell(row=next_row, column=4, value=web_price)
        sheet.cell(row=next_row, column=5, value=("yes" if product["PP1"] else "no"))
        sheet.cell(row=next_row, column=6, value=("yes" if product["PP2"] else "no"))
        sheet.cell(row=next_row, column=7, value=("yes" if product["PP3"] else "no"))
        sheet.cell(row=next_row, column=8, value=("yes" if product["PP4"] else "no"))
        sheet.cell(row=next_row, column=9, value=("yes" if product["PP5"] else "no"))
        sheet.cell(row=next_row, column=10, value=product["preorders"])
        workbook.save(file_path)
        print(f'Kaspi-Bot обновил цену: Новая: {web_price}, Старая: {stored_price}, Min Price: {min_price}')
